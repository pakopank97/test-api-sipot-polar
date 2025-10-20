"""
Pre-Validador SIPOT (Polars + JSON extendido + ACUSE DE ERRORES)
----------------------------------------------------------------
âœ… Agrupa errores contiguos por columna
âœ… Genera PDF institucional (logo grande + encabezado profesional)
âœ… Logs diarios (validacion_YYYY-MM-DD.log)
âœ… Genera JSON cuando no hay errores y lo expone vÃ­a /download/<filename>
"""

# ===============================================================
# IMPORTS
# ===============================================================
import polars as pl
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import json, re, os, uuid, threading, tempfile, time, logging, math
from logging.handlers import TimedRotatingFileHandler
from openpyxl import load_workbook
from dateutil import parser
from datetime import datetime
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from PIL import Image as PILImage

# ===============================================================
# CONFIG FLASK Y CARPETAS
# ===============================================================
app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'temp_uploads'
DOWNLOAD_FOLDER = 'temp_downloads'
LOG_FOLDER = 'logs'
STATIC_FOLDER = 'static'

for folder in (UPLOAD_FOLDER, DOWNLOAD_FOLDER, LOG_FOLDER, STATIC_FOLDER):
    os.makedirs(folder, exist_ok=True)

# ===============================================================
# CONFIGURACIÃ“N DE LOG (Windows y Linux)
# ===============================================================
fecha_actual = datetime.now().strftime("%Y-%m-%d")

# --- ðŸ§© OPCIÃ“N WINDOWS (Desarrollo local)
log_path = os.path.join(LOG_FOLDER, f"validacion_{fecha_actual}.log")

# --- ðŸ§© OPCIÃ“N LINUX (Servidor QA)
# log_path = f"/app/logs/validacion_{fecha_actual}.log"

handler = TimedRotatingFileHandler(
    log_path, when="midnight", interval=1, backupCount=30, encoding='utf-8', delay=True
)

def rotador_por_dia(name):
    base = os.path.splitext(name)[0]
    return f"{base}_{datetime.now().strftime('%Y-%m-%d')}.log"
handler.namer = rotador_por_dia

formatter = logging.Formatter(
    "%(asctime)s | %(levelname)s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
)
handler.setFormatter(formatter)

logger = logging.getLogger("validador")
logger.setLevel(logging.INFO)
if not logger.handlers:
    logger.addHandler(handler)

tasks = {}

# ===============================================================
# FUNCIONES AUXILIARES
# ===============================================================
def obtener_logo():
    logo_path = os.path.join(STATIC_FOLDER, "Logo_del_Gobierno_de_MÃ©xico.png")
    return logo_path if os.path.exists(logo_path) else None

def es_numero(v):
    try:
        float(v); return True
    except (ValueError, TypeError):
        return False

def es_fecha(v):
    try:
        parser.parse(str(v)); return True
    except Exception:
        return False

def es_hora(v):
    return bool(re.match(r'^([01]\d|2[0-3]):([0-5]\d)(:([0-5]\d))?$', str(v).strip()))

def es_url(v):
    return str(v).strip().lower().startswith(('http://', 'https://'))

def es_anio(v):
    return es_numero(v) and len(str(v).split('.')[0]) == 4

def esta_vacio(v):
    if v is None:
        return True
    if isinstance(v, (int, float)):
        if isinstance(v, float) and math.isnan(v):
            return True
        return False
    s = str(v).strip()
    if s in ('0', '0.0'):
        return False
    return s == '' or s.lower() in ('nan', 'none', 'null')

VALIDADORES = {
    '3': {'func': es_numero, 'nombre': 'NÃºmero'},
    '4': {'func': es_fecha,  'nombre': 'Fecha'},
    '5': {'func': es_hora,   'nombre': 'Hora (HH:MM)'},
    '6': {'func': es_numero, 'nombre': 'Moneda'},
    '7': {'func': es_url,    'nombre': 'URL'},
    '12': {'func': es_anio,  'nombre': 'AÃ±o (4 dÃ­gitos)'},
    '13': {'func': es_fecha, 'nombre': 'Fecha'},
}

def obtener_coordenada_excel(fila, col):
    col_str = ""
    while col >= 0:
        col_str = chr(ord('A') + col % 26) + col_str
        col = col // 26 - 1
    return f"{col_str}{fila + 1}"

# ===============================================================
# CONVERSIÃ“N EXCEL â†’ CSV
# ===============================================================
def convertir_excel_a_csv(ruta_excel):
    temp_csv = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active
    with open(temp_csv.name, "w", encoding="utf-8") as f:
        for row in ws.iter_rows(values_only=True):
            vals = []
            for c in row:
                if c == 0 or c == 0.0 or c == "0":
                    vals.append("0")
                elif c is None:
                    vals.append("")
                else:
                    vals.append(str(c).replace(",", " ").replace("\n", " ").replace("\r", " "))
            f.write(",".join(vals) + "\n")
    return temp_csv.name

# ===============================================================
# PROCESAMIENTO PRINCIPAL (con bloques de errores y JSON en Ã©xito)
# ===============================================================
def procesar_archivo_en_segundo_plano(filepath, task_id):
    t0 = time.time()
    csv_path = None
    try:
        ext = os.path.splitext(filepath)[1].lower()
        csv_path = convertir_excel_a_csv(filepath) if ext in (".xlsx", ".xls") else filepath

        with open(csv_path, 'r', encoding='utf-8') as f:
            first = f.readline()
            n_cols = len(first.strip().split(',')) if first else 0
        schema = {str(i): pl.Utf8 for i in range(n_cols)}

        df = pl.read_csv(csv_path, has_header=False, infer_schema_length=0,
                         null_values=['', 'NULL', 'null', 'NaN', 'nan'],
                         schema_overrides=schema)

        # eliminar filas completamente vacÃ­as
        df = df.filter(pl.any_horizontal(~pl.col("*").is_null() & (pl.col("*").cast(str).str.strip_chars() != "")))

        reglas = df.row(3) if df.height > 3 else []
        headers_visibles = [str(h or "").strip() for h in (df.row(6) if df.height > 6 else [])]
        datos = df.slice(7)
        lista_de_errores = []

        # ValidaciÃ³n base
        for fila_idx, row in enumerate(datos.iter_rows()):
            abs_row_idx = fila_idx + 7
            for col_idx, valor in enumerate(row):
                if col_idx >= len(headers_visibles):
                    continue
                header = headers_visibles[col_idx]
                if header == '':
                    continue
                if esta_vacio(valor):
                    coord = obtener_coordenada_excel(abs_row_idx, col_idx)
                    lista_de_errores.append(f"Celda {coord} bajo '{header}' vacÃ­a.")
                else:
                    regla = str(reglas[col_idx]).split('.')[0] if col_idx < len(reglas) else '0'
                    if regla in VALIDADORES:
                        val = VALIDADORES[regla]
                        if not val['func'](valor):
                            coord = obtener_coordenada_excel(abs_row_idx, col_idx)
                            lista_de_errores.append(f"Celda {coord} ('{valor}') invÃ¡lida. Se esperaba: {val['nombre']}.")

        # Agrupar errores contiguos
        if lista_de_errores:
            patron = re.compile(r"Celda\s+([A-Z]+)(\d+)\s+(.*)")
            por_col_y_msg = {}
            for err in lista_de_errores:
                m = patron.search(err)
                if not m:
                    continue
                col, fila, msg = m.group(1), int(m.group(2)), m.group(3).strip()
                por_col_y_msg.setdefault((col, msg), []).append(fila)

            bloques = []
            for (col, msg), filas in por_col_y_msg.items():
                filas.sort()
                ini, prev = filas[0], filas[0]
                for f in filas[1:]:
                    if f == prev + 1:
                        prev = f
                        continue
                    bloques.append(f"Celda {col}{ini}" + (f" hasta {col}{prev}" if ini != prev else "") + f" {msg}")
                    ini = prev = f
                bloques.append(f"Celda {col}{ini}" + (f" hasta {col}{prev}" if ini != prev else "") + f" {msg}")
            lista_de_errores = bloques

        # Metadata para logging/resultado
        nombre_corto = df[2, 3] if df.height > 2 and df.width > 3 else 'N/D'
        size_mb = os.path.getsize(filepath) / (1024 * 1024)
        dur = round(time.time() - t0, 1)
        n_err = len(lista_de_errores)

        # Si hay errores -> solo devolver errores
        if lista_de_errores:
            logger.info(f"[{task_id}] Archivo: {os.path.basename(filepath)} | Nombre Corto: {nombre_corto} | TamaÃ±o: {size_mb:.2f} MB | Errores: {n_err} | Tiempo: {dur:.1f}s | Estado: ERROR")
            tasks[task_id] = {
                'status': 'complete',
                'result': {
                    'status': 'error',
                    'errors': lista_de_errores,
                    'nombre_corto': str(nombre_corto)
                }
            }
            return

        # Si NO hay errores -> construir JSON y exponer nombre de archivo
        id_formato = df[0, 0] if df.height > 0 else "Formato no encontrado"
        titulo = df[2, 0] if df.height > 2 else ""
        headers_backend = [str(h or "header_sin_nombre").strip() for h in (df.row(6) if df.height > 6 else [])]
        datos_backend = df.slice(7)

        registros = []
        for r in datos_backend.iter_rows():
            reg = {}
            for idx, val in enumerate(r):
                if idx < len(headers_backend):
                    reg[headers_backend[idx]] = str(val or "")
            registros.append(reg)

        out_json = {
            "id_formato": str(id_formato),
            "Titulo": str(titulo).strip(),
            "Nombre Corto": str(nombre_corto).strip(),
            "data": registros
        }

        json_filename = f"{task_id}.json"
        json_path = os.path.join(DOWNLOAD_FOLDER, json_filename)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(out_json, f, ensure_ascii=False, indent=2)

        logger.info(f"[{task_id}] Archivo: {os.path.basename(filepath)} | Nombre Corto: {nombre_corto or 'N/D'} | TamaÃ±o: {size_mb:.2f} MB | Errores: 0 | Tiempo: {dur:.1f}s | Estado: OK")

        tasks[task_id] = {
            'status': 'complete',
            'result': {
                'status': 'success',
                'download_file': json_filename,
                'nombre_corto': str(nombre_corto)
            }
        }

    except Exception as e:
        logger.error(f"[{task_id}] Error inesperado: {str(e)}")
        tasks[task_id] = {'status': 'failed', 'error': str(e)}
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)
        if csv_path and csv_path != filepath and os.path.exists(csv_path):
            os.remove(csv_path)

# ===============================================================
# ENDPOINTS
# ===============================================================
@app.route('/')
def home():
    return send_from_directory('.', 'index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'archivo' not in request.files:
        return jsonify({"error": "No se encontrÃ³ archivo"}), 400
    archivo = request.files['archivo']
    if archivo.filename == '':
        return jsonify({"error": "No se seleccionÃ³ archivo"}), 400
    ruta = os.path.join(UPLOAD_FOLDER, f"{uuid.uuid4()}_{archivo.filename}")
    archivo.save(ruta)
    task_id = str(uuid.uuid4())
    tasks[task_id] = {'status': 'processing'}
    threading.Thread(target=procesar_archivo_en_segundo_plano, args=(ruta, task_id)).start()
    return jsonify({'task_id': task_id})

@app.route('/status/<task_id>')
def status(task_id):
    return jsonify(tasks.get(task_id, {'status': 'not_found'}))

@app.route('/download/<filename>')
def download(filename):
    return send_from_directory(DOWNLOAD_FOLDER, filename, as_attachment=True)

# ===============================================================
# MAIN
# ===============================================================
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8081)
# ===============================================================